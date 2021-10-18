print('Extractor importado')

## MÒDULOS
from pathlib import Path
import openpyxl
import os
import re



def Ingresar(Nombre_del_archivo, Archivo_colegiados):

## Abro Archivos
    BASE_DIR = str(Path().resolve())
    MEDIA_FILE = BASE_DIR + '/media/'
    libro_1 = openpyxl.load_workbook(MEDIA_FILE + Nombre_del_archivo)
    libro_2 = openpyxl.load_workbook(MEDIA_FILE + Archivo_colegiados)
    hoja_1 = libro_1.active
    hoja_2 = libro_2.active
    print('Archivos abiertos')
    cantidad_de_filas_1 = hoja_1.max_row
    cantidad_de_filas_2 = hoja_2.max_row
## Relevo el campo de pagos, si es mayor a 0 extraigo cuit_cuil
    for x in range (2, cantidad_de_filas_1):
        if hoja_1.cell(row=x,column=6).value == None:
            pass
        elif hoja_1.cell(row=x,column=6).value > 0:
            celda = hoja_1.cell(row=x,column=3).value
            cuit_cuil = ("")
            for i in celda:
                if i.isdigit():
                    cuit_cuil += i             
            hoja_1.cell(row=x,column=9).value = cuit_cuil
# Comparo el número extraído con el CUIT de la planilla de colegiados.
# Si coinciden copio el número y nombre del colegiado.
            for i in range(2, cantidad_de_filas_2):
                primer_número = hoja_1.cell(row = x, column = 9).value
                segundo_número = hoja_2.cell(row = i, column = 2).value
                if primer_número == segundo_número:
                    hoja_1.cell(row = x, column = 10).value = hoja_2.cell(row = i, column = 1).value
                    hoja_1.cell(row = x, column = 11).value = hoja_2.cell(row = i, column = 3).value
# Grabo Modificaciones
    libro_1.save(filename = MEDIA_FILE + Nombre_del_archivo)
    print('proceso terminado')

